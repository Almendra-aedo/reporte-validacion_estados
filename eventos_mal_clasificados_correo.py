import html
import mimetypes
import os
import smtplib
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================
ZONA_HORARIA = ZoneInfo("America/Santiago")

URL_VEHICULOS = "https://external.driv.in/api/external/v2/vehicles"
URL_EVENTOS = (
    "https://external.driv.in/api/external/v2/"
    "schedulable_events/events_abastible"
)

ARCHIVO_EXCEL = Path(
    os.getenv("ARCHIVO_EXCEL", "validacion_estado_flota.xlsx")
)
ARCHIVO_HTML = Path(
    os.getenv("ARCHIVO_HTML", "reporte_validacion_flota.html")
)
ARCHIVO_ASUNTO = Path(
    os.getenv("ARCHIVO_ASUNTO", "asunto_correo.txt")
)

PREFIJO_ASUNTO = os.getenv(
    "PREFIJO_ASUNTO",
    "[VALIDACION FLOTA]",
).strip()

COLUMNAS_EVENTO = [
    "vehicle_code",
    "description",
    "start_date",
    "end_date",
    "correlative",
]

COLUMNAS_CORREO = [
    "Código vehículo",
    "Dependencia",
    "Modelo",
    "Descripción vehículo",
    "Estado API",
    "Tiene evento abierto",
    "Motivo",
    "Descripción eventos abiertos",
    "Fecha inicio eventos",
]


# ============================================================
# UTILIDADES DE CONFIGURACIÓN
# ============================================================
def obtener_variable_obligatoria(nombre):
    valor = os.getenv(nombre, "").strip()
    if not valor:
        raise RuntimeError(
            f"Falta configurar la variable de entorno obligatoria: {nombre}"
        )
    return valor


def leer_booleano(nombre, valor_por_defecto=False):
    valor = os.getenv(nombre)
    if valor is None or not valor.strip():
        return valor_por_defecto

    return valor.strip().lower() in {
        "1",
        "true",
        "sí",
        "si",
        "yes",
        "y",
        "on",
    }


def leer_entero(nombre, valor_por_defecto):
    valor = os.getenv(nombre, "").strip()
    return int(valor) if valor else valor_por_defecto


def separar_correos(texto):
    if not texto:
        return []

    texto = texto.replace(";", ",")
    return [
        correo.strip()
        for correo in texto.split(",")
        if correo.strip()
    ]


# ============================================================
# CONSULTA DE APIS
# ============================================================
def consultar(url, params=None):
    api_key = obtener_variable_obligatoria("DRIVIN_API_KEY")

    respuesta = requests.get(
        url,
        headers={
            "x-api-key": api_key,
            "Accept": "application/json",
        },
        params=params,
        timeout=90,
    )
    respuesta.raise_for_status()

    datos = respuesta.json()
    if isinstance(datos, dict):
        return datos.get("response", [])
    return datos


def asegurar_columnas(dataframe, columnas):
    dataframe = dataframe.copy()
    for columna in columnas:
        if columna not in dataframe.columns:
            dataframe[columna] = pd.NA
    return dataframe


def es_vacio(serie):
    texto = serie.astype(str).str.strip().str.lower()
    return serie.isna() | texto.isin(
        {"", "none", "nan", "nat", "null"}
    )


def obtener_dependencia(fleets):
    if isinstance(fleets, list):
        nombre = str(fleets[0]).strip() if fleets else ""
    else:
        nombre = str(fleets or "").split(",")[0].strip()

    if nombre.lower().startswith("flota "):
        nombre = nombre[6:].strip()

    return nombre or "Sin dependencia"


def unir_valores(serie):
    valores = {
        str(valor).strip()
        for valor in serie
        if pd.notna(valor) and str(valor).strip()
    }
    return " | ".join(sorted(valores))


# ============================================================
# EXCEL
# ============================================================
def ajustar_hoja(hoja):
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    relleno_encabezado = PatternFill("solid", fgColor="0B7185")
    relleno_ok = PatternFill("solid", fgColor="E2F0D9")
    relleno_nok = PatternFill("solid", fgColor="FCE4D6")

    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    encabezados = {
        celda.value: celda.column
        for celda in hoja[1]
    }

    columna_validacion = encabezados.get("Validación")

    if columna_validacion:
        for fila in range(2, hoja.max_row + 1):
            valor = hoja.cell(fila, columna_validacion).value
            relleno = relleno_ok if valor == "OK" else relleno_nok

            for celda in hoja[fila]:
                celda.fill = relleno
                celda.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    for columna in hoja.columns:
        letra = columna[0].column_letter
        ancho = max(
            len(str(celda.value or ""))
            for celda in columna
        ) + 2

        hoja.column_dimensions[letra].width = min(
            max(ancho, 12),
            45,
        )


def guardar_excel(resultado, solo_nok, resumen):
    with pd.ExcelWriter(
        ARCHIVO_EXCEL,
        engine="openpyxl",
    ) as writer:
        resultado.to_excel(
            writer,
            sheet_name="Validación completa",
            index=False,
        )

        solo_nok.to_excel(
            writer,
            sheet_name="Solo NOK",
            index=False,
        )

        resumen.to_excel(
            writer,
            sheet_name="Resumen",
            index=False,
        )

        ajustar_hoja(writer.sheets["Validación completa"])
        ajustar_hoja(writer.sheets["Solo NOK"])
        ajustar_hoja(writer.sheets["Resumen"])


# ============================================================
# HTML DEL CORREO
# ============================================================
def mostrar_valor(valor):
    if pd.isna(valor):
        return "—"

    texto = str(valor).strip()
    if texto.lower() in {"", "nan", "nat", "none", "null"}:
        return "—"

    return texto


def formatear_fecha_para_correo(valor):
    texto = mostrar_valor(valor)
    if texto == "—":
        return texto

    fechas_formateadas = []
    for parte in texto.split(" | "):
        parte = parte.strip()
        fecha = pd.to_datetime(parte, errors="coerce")

        if pd.isna(fecha):
            fechas_formateadas.append(parte)
        else:
            fechas_formateadas.append(
                fecha.strftime("%d-%m-%Y %H:%M")
            )

    return " | ".join(fechas_formateadas)


def crear_tarjeta_kpi(titulo, valor, fondo, borde):
    return f"""
    <td width="50%" style="padding:6px;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0"
             style="border-collapse:collapse;background:{fondo};border:1px solid {borde};border-radius:8px;">
        <tr>
          <td style="padding:14px 16px 4px 16px;font-family:Arial,sans-serif;font-size:13px;color:#475569;">
            {html.escape(titulo)}
          </td>
        </tr>
        <tr>
          <td style="padding:0 16px 14px 16px;font-family:Arial,sans-serif;font-size:28px;font-weight:bold;color:#0f172a;">
            {html.escape(str(valor))}
          </td>
        </tr>
      </table>
    </td>
    """


def crear_tabla_html(solo_nok):
    if solo_nok.empty:
        return """
        <div style="margin-top:18px;padding:16px;border:1px solid #70AD47;background:#E2F0D9;border-radius:8px;font-family:Arial,sans-serif;color:#1f5132;">
          No se detectaron vehículos incorrectamente clasificados.
        </div>
        """

    tabla = solo_nok[COLUMNAS_CORREO].copy()
    tabla["Fecha inicio eventos"] = tabla[
        "Fecha inicio eventos"
    ].apply(formatear_fecha_para_correo)

    encabezado = "".join(
        f"<th style='padding:9px 8px;border:1px solid #d1d5db;background:#0B7185;color:#ffffff;font-family:Arial,sans-serif;font-size:11px;text-align:left;vertical-align:top;'>{html.escape(columna)}</th>"
        for columna in COLUMNAS_CORREO
    )

    filas = []
    for _, fila in tabla.iterrows():
        celdas = []

        for columna in COLUMNAS_CORREO:
            valor = mostrar_valor(fila[columna])
            valor_seguro = html.escape(valor).replace("\n", "<br>")

            fondo = "#ffffff"
            if columna == "Motivo":
                fondo = "#fff2cc"
            elif columna == "Estado API":
                fondo = (
                    "#e2f0d9"
                    if valor == "Activo"
                    else "#fce4d6"
                )

            celdas.append(
                "<td style='"
                f"padding:8px;border:1px solid #d1d5db;background:{fondo};"
                "font-family:Arial,sans-serif;font-size:11px;color:#1f2937;"
                "vertical-align:top;line-height:1.35;'>"
                f"{valor_seguro}</td>"
            )

        filas.append(f"<tr>{''.join(celdas)}</tr>")

    return f"""
    <div style="margin-top:22px;overflow-x:auto;">
      <table role="table" cellspacing="0" cellpadding="0"
             style="border-collapse:collapse;width:100%;min-width:1100px;">
        <thead><tr>{encabezado}</tr></thead>
        <tbody>{''.join(filas)}</tbody>
      </table>
    </div>
    """


def crear_html_correo(solo_nok, kpis, fecha_hora):
    fecha_visible = fecha_hora.strftime("%d-%m-%Y %H:%M")

    fila_1 = (
        crear_tarjeta_kpi(
            "Correctamente clasificados",
            kpis["Correctamente clasificados"],
            "#E2F0D9",
            "#70AD47",
        )
        + crear_tarjeta_kpi(
            "Incorrectamente clasificados",
            kpis["Incorrectamente clasificados"],
            "#FCE4D6",
            "#C00000",
        )
    )

    fila_2 = (
        crear_tarjeta_kpi(
            "Inactivos sin evento abierto",
            kpis["Inactivos sin evento abierto"],
            "#FFF2CC",
            "#BF9000",
        )
        + crear_tarjeta_kpi(
            "Activos con evento abierto",
            kpis["Activos con evento abierto"],
            "#FCE4D6",
            "#ED7D31",
        )
    )

    tabla_html = crear_tabla_html(solo_nok)

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Validación de estado de flota</title>
</head>
<body style="margin:0;padding:0;background:#f3f4f6;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;background:#f3f4f6;">
    <tr>
      <td align="center" style="padding:24px 12px;">
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0"
               style="max-width:1250px;border-collapse:collapse;background:#ffffff;border:1px solid #e5e7eb;border-radius:10px;">
          <tr>
            <td style="padding:24px 24px 8px 24px;">
              <div style="font-family:Arial,sans-serif;font-size:24px;font-weight:bold;color:#0f172a;">
                Validación de estado de flota
              </div>
              <div style="margin-top:6px;font-family:Arial,sans-serif;font-size:13px;color:#64748b;">
                Fecha de ejecución: {html.escape(fecha_visible)} · Vehículos revisados: {html.escape(str(kpis['Vehículos revisados']))}
              </div>
            </td>
          </tr>
          <tr>
            <td style="padding:8px 18px 0 18px;">
              <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;">
                <tr>{fila_1}</tr>
                <tr>{fila_2}</tr>
              </table>
            </td>
          </tr>
          <tr>
            <td style="padding:8px 24px 24px 24px;">
              <div style="margin-top:14px;font-family:Arial,sans-serif;font-size:17px;font-weight:bold;color:#0f172a;">
                Vehículos incorrectamente clasificados
              </div>
              <div style="margin-top:4px;font-family:Arial,sans-serif;font-size:12px;color:#64748b;">
                Se muestran únicamente los registros NOK. El detalle completo se adjunta en Excel.
              </div>
              {tabla_html}
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>
"""


# ============================================================
# ENVÍO DEL CORREO
# ============================================================
def adjuntar_archivo(mensaje, ruta):
    tipo_mime, _ = mimetypes.guess_type(ruta.name)

    if tipo_mime:
        tipo_principal, subtipo = tipo_mime.split("/", 1)
    else:
        tipo_principal, subtipo = (
            "application",
            "octet-stream",
        )

    mensaje.add_attachment(
        ruta.read_bytes(),
        maintype=tipo_principal,
        subtype=subtipo,
        filename=ruta.name,
    )


def enviar_correo(asunto, cuerpo_html):
    smtp_host = obtener_variable_obligatoria("SMTP_HOST")
    smtp_port = leer_entero("SMTP_PORT", 587)
    smtp_user = obtener_variable_obligatoria("SMTP_USER")
    smtp_password = obtener_variable_obligatoria("SMTP_PASSWORD")

    remitente = os.getenv("EMAIL_FROM", "").strip() or smtp_user
    destinatarios = separar_correos(
        obtener_variable_obligatoria("EMAIL_TO")
    )
    copias = separar_correos(os.getenv("EMAIL_CC", ""))

    usar_ssl = leer_booleano("SMTP_USE_SSL", Tue)
    usar_starttls = leer_booleano(
        "SMTP_USE_STARTTLS",
        not usar_ssl,
    )

    mensaje = EmailMessage()
    mensaje["Subject"] = asunto
    mensaje["From"] = remitente
    mensaje["To"] = ", ".join(destinatarios)

    if copias:
        mensaje["Cc"] = ", ".join(copias)

    mensaje.set_content(
        "Reporte de validación de estado de flota. "
        "El correo requiere visualización HTML."
    )
    mensaje.add_alternative(cuerpo_html, subtype="html")

    adjuntar_archivo(mensaje, ARCHIVO_EXCEL)

    receptores = destinatarios + copias

    if usar_ssl:
        with smtplib.SMTP_SSL(
            smtp_host,
            smtp_port,
            timeout=90,
        ) as servidor:
            servidor.login(smtp_user, smtp_password)
            servidor.send_message(
                mensaje,
                from_addr=remitente,
                to_addrs=receptores,
            )
    else:
        with smtplib.SMTP(
            smtp_host,
            smtp_port,
            timeout=90,
        ) as servidor:
            servidor.ehlo()

            if usar_starttls:
                servidor.starttls()
                servidor.ehlo()

            servidor.login(smtp_user, smtp_password)
            servidor.send_message(
                mensaje,
                from_addr=remitente,
                to_addrs=receptores,
            )


# ============================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================
def generar_reporte():
    fecha_hora = datetime.now(ZONA_HORARIA)
    fecha_api = fecha_hora.strftime("%Y-%m-%d")

    datos_vehiculos = consultar(URL_VEHICULOS)
    datos_eventos = consultar(
        URL_EVENTOS,
        {
            "from_datetime": fecha_api,
            "is_unfinished_event": 1,
            "type": "vehicle",
        },
    )

    vehiculos = pd.DataFrame(datos_vehiculos)
    vehiculos = asegurar_columnas(
        vehiculos,
        ["code", "fleets", "is_active", "model", "description"],
    )
    vehiculos = vehiculos[
        ["code", "fleets", "is_active", "model", "description"]
    ].copy()

    eventos = pd.json_normalize(
        datos_eventos,
        sep=".",
    )
    eventos = asegurar_columnas(eventos, COLUMNAS_EVENTO)

    vehiculos["code"] = (
        vehiculos["code"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    def normalizar_fleets(valor):
        if isinstance(valor, list):
            return valor
        if pd.isna(valor):
            return ""
        return str(valor).strip()

    vehiculos["fleets"] = vehiculos["fleets"].apply(
        normalizar_fleets
    )

    vehiculos = vehiculos[
        ~vehiculos["code"].str.contains("FI", na=False)
        & vehiculos["code"].str.len().le(6)
        & ~vehiculos["fleets"].astype(str).str.lower().str.contains(
            "emergencia",
            na=False,
        )
    ].copy()

    vehiculos["Dependencia"] = vehiculos["fleets"].apply(
        obtener_dependencia
    )

    if not eventos.empty:
        eventos = eventos[
            ~es_vacio(eventos["start_date"])
            & es_vacio(eventos["end_date"])
        ].copy()

    eventos["vehicle_code"] = (
        eventos["vehicle_code"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    if eventos.empty:
        resumen_eventos = pd.DataFrame(
            columns=[
                "vehicle_code",
                "Cantidad_eventos_abiertos",
                "Eventos",
                "Fecha_inicio_evento",
                "Correlativos",
            ]
        )
    else:
        resumen_eventos = (
            eventos.groupby("vehicle_code", dropna=False)
            .agg(
                Cantidad_eventos_abiertos=(
                    "vehicle_code",
                    "size",
                ),
                Eventos=("description", unir_valores),
                Fecha_inicio_evento=("start_date", unir_valores),
                Correlativos=("correlative", unir_valores),
            )
            .reset_index()
        )

    resultado = vehiculos.merge(
        resumen_eventos,
        left_on="code",
        right_on="vehicle_code",
        how="left",
    )

    resultado["Cantidad_eventos_abiertos"] = (
        resultado["Cantidad_eventos_abiertos"]
        .fillna(0)
        .astype(int)
    )

    resultado["Tiene evento abierto"] = (
        resultado["Cantidad_eventos_abiertos"]
        .gt(0)
        .map({True: "Sí", False: "No"})
    )

    resultado["Estado API"] = resultado["is_active"].map(
        {
            True: "Activo",
            False: "Inactivo",
        }
    ).fillna("Sin estado")

    resultado["Validación"] = (
        (
            resultado["is_active"].eq(False)
            & resultado["Cantidad_eventos_abiertos"].gt(0)
        )
        |
        (
            resultado["is_active"].eq(True)
            & resultado["Cantidad_eventos_abiertos"].eq(0)
        )
    ).map(
        {
            True: "OK",
            False: "NOK",
        }
    )

    resultado["Motivo"] = ""

    resultado.loc[
        resultado["is_active"].eq(False)
        & resultado["Cantidad_eventos_abiertos"].eq(0),
        "Motivo",
    ] = "Vehículo inactivo sin evento abierto"

    resultado.loc[
        resultado["is_active"].eq(True)
        & resultado["Cantidad_eventos_abiertos"].gt(0),
        "Motivo",
    ] = "Vehículo activo con evento abierto"

    resultado = resultado.rename(
        columns={
            "code": "Código vehículo",
            "model": "Modelo",
            "description": "Descripción vehículo",
            "Eventos": "Descripción eventos abiertos",
            "Fecha_inicio_evento": "Fecha inicio eventos",
            "Correlativos": "Correlativos eventos",
            "Cantidad_eventos_abiertos": "Cantidad eventos abiertos",
        }
    )

    columnas_finales = [
        "Código vehículo",
        "Dependencia",
        "Modelo",
        "Descripción vehículo",
        "Estado API",
        "Tiene evento abierto",
        "Cantidad eventos abiertos",
        "Validación",
        "Motivo",
        "Descripción eventos abiertos",
        "Fecha inicio eventos",
        "Correlativos eventos",
    ]

    resultado = (
        resultado[columnas_finales]
        .sort_values(
            ["Validación", "Dependencia", "Código vehículo"],
            ascending=[False, True, True],
        )
        .reset_index(drop=True)
    )

    solo_nok = resultado[
        resultado["Validación"].eq("NOK")
    ].copy()

    kpis = {
        "Vehículos revisados": int(len(resultado)),
        "Correctamente clasificados": int(
            resultado["Validación"].eq("OK").sum()
        ),
        "Incorrectamente clasificados": int(
            resultado["Validación"].eq("NOK").sum()
        ),
        "Inactivos sin evento abierto": int(
            resultado["Motivo"].eq(
                "Vehículo inactivo sin evento abierto"
            ).sum()
        ),
        "Activos con evento abierto": int(
            resultado["Motivo"].eq(
                "Vehículo activo con evento abierto"
            ).sum()
        ),
    }

    resumen = pd.DataFrame(
        {
            "Indicador": [
                "Fecha consultada",
                "Hora de ejecución",
                *kpis.keys(),
            ],
            "Valor": [
                fecha_api,
                fecha_hora.strftime("%H:%M:%S"),
                *kpis.values(),
            ],
        }
    )

    guardar_excel(resultado, solo_nok, resumen)

    cuerpo_html = crear_html_correo(
        solo_nok,
        kpis,
        fecha_hora,
    )
    ARCHIVO_HTML.write_text(cuerpo_html, encoding="utf-8")

    asunto = (
        f"{PREFIJO_ASUNTO} Reporte estado de flota | "
        f"{fecha_hora.strftime('%d-%m-%Y %H:%M')}"
    )
    ARCHIVO_ASUNTO.write_text(asunto, encoding="utf-8")

    github_output = os.getenv("GITHUB_OUTPUT", "").strip()
    if github_output:
        with open(github_output, "a", encoding="utf-8") as salida:
            salida.write(f"asunto={asunto}\n")
            salida.write(f"archivo_excel={ARCHIVO_EXCEL}\n")
            salida.write(f"archivo_html={ARCHIVO_HTML}\n")

    print(f"Reporte Excel generado: {ARCHIVO_EXCEL}")
    print(f"Vista HTML generada: {ARCHIVO_HTML}")
    print(f"Vehículos revisados: {kpis['Vehículos revisados']}")
    print(
        "Vehículos NOK: "
        f"{kpis['Incorrectamente clasificados']}"
    )

    if leer_booleano("ENVIAR_CORREO", False):
        enviar_correo(asunto, cuerpo_html)
        print("Correo enviado correctamente.")
    else:
        print(
            "El correo no se envió porque ENVIAR_CORREO "
            "no está configurado como true."
        )


if __name__ == "__main__":
    generar_reporte()
